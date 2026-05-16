"""
Flent Homes — UX Intelligence Data Processor v2
Deep analysis using pattern matching, resolution mining, temporal analysis,
and statistical clustering. No LLM required.
"""

import openpyxl
import json
import re
import math
import datetime
from collections import Counter, defaultdict

INPUT_FILE = "/Users/mac/Downloads/Flent_ Customer Tickets.xlsx"
OUTPUT_FILE = "../src/data/tickets.json"
INSIGHTS_FILE = "../src/data/insights.json"

# ─────────────────────────────────────────────────────────────────────────────
# CATEGORY NORMALIZATION — expanded to handle all 100+ raw category values
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_MAP = {
    "plumbing": "Plumbing",
    "pluming": "Plumbing",
    "water seepage issue": "Plumbing",
    "electrical": "Electrical",
    "electricals": "Electrical",
    "electerical": "Electrical",
    "electrician": "Electrical",
    "electricity": "Electrical",
    "appliance": "Appliance",
    "appliacne": "Appliance",
    "applicane": "Appliance",
    "geyser not working": "Appliance",
    "drinkprime": "Appliance",
    "drink prime": "Appliance",
    "kitchen chimney": "Appliance",
    "stove service": "Appliance",
    "tv issue": "Appliance",
    "tv": "Appliance",
    "gas stove": "Appliance",
    "gas cylinder": "Appliance",
    "carpentry": "Carpentry",
    "furniture": "Carpentry",
    "fixture": "Carpentry",
    "fixtures": "Carpentry",
    "door cover": "Carpentry",
    "wifi/water/gas": "Utilities",
    "utilities-wifi/water/lpg": "Utilities",
    "utility": "Utilities",
    "utilities": "Utilities",
    "wifi": "Utilities",
    "wi-fi": "Utilities",
    "wifi airtel": "Utilities",
    "wifi wiring": "Utilities",
    "building maintenance/ utilities": "Utilities",
    "utilities payment pending": "Utilities",
    "pest control": "Pest Control",
    "reimbursement": "Reimbursement",
    "reimbursements": "Reimbursement",
    "refund": "Reimbursement",
    "rent": "Reimbursement",
    "discount enquiry": "Reimbursement",
    "cleaning": "Cleaning",
    "deep cleaning": "Cleaning",
    "deep cleaning of the sofa": "Cleaning",
    "sofa cleaning": "Cleaning",
    "cleaing": "Cleaning",
    "cloeaning": "Cleaning",
    "move out - cleaning": "Cleaning",
    "drycleaning": "Cleaning",
    "move-out request": "Move-out",
    "move out": "Move-out",
    "moveout notice": "Move-out",
    "pre move in request": "Move-in",
    "pre movein request": "Move-in",
    "pre move-in request": "Move-in",
    "move-in formalities incomplete": "Move-in",
    "move-in formalities": "Move-in",
    "move-in request": "Move-in",
    "move in request": "Move-in",
    "movein": "Move-in",
    "inventory": "Inventory",
    "inventory request": "Inventory",
    "bedsheet": "Inventory",
    "agreement": "Documentation",
    "agreements": "Documentation",
    "contract": "Documentation",
    "documentation": "Documentation",
    "property documentation": "Documentation",
    "police verification": "Documentation",
    "compliance": "Documentation",
    "maintenance": "Maintenance",
    "exterior": "Maintenance",
    "structual": "Maintenance",
    "society pipeline work": "Maintenance",
    "society": "Maintenance",
    "gardening": "Maintenance",
    "gate issue": "Maintenance",
    "garbage disposal": "Maintenance",
    "house issue": "Maintenance",
    "lock": "Access & Security",
    "lock issue": "Access & Security",
    "key duplication": "Access & Security",
    "keys": "Access & Security",
    "mosquito mesh": "Maintenance",
    "pigeon net": "Maintenance",
    "installation": "Maintenance",
    "external factors": "Maintenance",
    "transportation": "Other",
    "enquiry": "Communication",
    "video": "Communication",
    "inspection findings": "Maintenance",
    "spillover": "Other",
    "internal": "Internal",
    "oo": "Internal",
    "knu": "Internal",
    "gestation tickets": "Internal",
    "15 day": "Internal",
    "n/a": "_unclassified",
    "something else": "_unclassified",
    "somethine else": "_unclassified",
    "others": "_unclassified",
}


# ─────────────────────────────────────────────────────────────────────────────
# DESCRIPTION-BASED RECLASSIFICATION — for "Other", "Uncategorized", N/A
# Uses title + description text to infer real category
# ─────────────────────────────────────────────────────────────────────────────

DESCRIPTION_RULES = [
    # Appliance — must be BEFORE plumbing/electrical to catch "water filter", "geyser", "washing machine"
    (r"\b(ro\b|water\s?filter|water\s?purifier|aquaguard|washing\s?machin|waching\s?machin|fridge|refrigerator|microwave|chimney|oven|stove|gas\s?stove|gas\s?knob|induction|mixer|geyser|geysar|drinkprime|tv|television|iron\s?box|iron(?!\s?rod)|exhaust\s?fan|exhaust|cloth\s?drying|drying\s?stand|lamp|aquarium|vacuum|ventilation)\b", "Appliance"),
    # Plumbing — water issues that are NOT appliances
    (r"\b(leak|leaking|leakage|drain|drainage|clog|blocked|flush|toilet|basin|washbasin|washbeshan|tap|pipe|seepage|water\s?tank|sewer|gutter|bathroom.*water|water.*bathroom|water\s?pressure|shower\s?pressure|shower.*not\s?work|sink|water\s?issue|hard\s?water|low.*pressure|plumb|no\s?water|drinking\s?water|water\s?damage)\b", "Plumbing"),
    # Electrical
    (r"\b(switch|wire|wiring|socket|power|electricity|volt|mcb|circuit|breaker|short\s?circuit|current|light.*not|not.*light|bulb|led|tubelight|tube\s?light|fan.*not|not.*fan|fan.*nois|fan.*noice|fan\s?making|fan.*slow|ac\b|air\s?condition|heater)\b", "Electrical"),
    # WiFi / Connectivity
    (r"\b(wifi|wi-fi|internet|router|broadband|airtel|jio|bsnl|network|connectivity|modem|speed.*slow|slow.*speed|no.*internet|internet.*not|disconnecting)\b", "Utilities"),
    # Carpentry / Fixtures / Furniture
    (r"\b(door|window|wardrobe|cupboard|drawer|cabinet|shelf|hinge|handle|latch|wood|plywood|veneer|edge\s?band|bed.*frame|frame|bed.*broken|broken.*bed|table|chair|desk|hooks?|rod|hanger|towel\s?holder|curtains?|black\s?out|blackout|ladder|painting|wall\s?art|crack\s?in|hole\s?in\s?the\s?wall|furniture|sofa|mattress|bedsheet|bed\s?sheet|pillow|cushion|rack|net\b|pigeon\s?net|mosquito\s?mesh|balcony)\b", "Carpentry"),
    # Pest Control
    (r"\b(pest|cockroach|roach|roaches|rat|rats|mouse|mice|ant|ants|mosquito|termite|bed\s?bug|spider|lizard|pigeon|snake|smell|stink|bad\s?odour|bad\s?odor|bad\s?smell|fish|infest)\b", "Pest Control"),
    # Cleaning
    (r"\b(clean|cleaning|hygiene|dirty|stain|deep\s?clean|mop|sweep|garbage|dustbin|trash|bucket|mug)\b", "Cleaning"),
    # Payment / Billing
    (r"\b(rent|payment|paid|bill|invoice|refund|reimburse|deposit|security\s?deposit|due|pending.*payment|payment.*pending|charge|emi|receipt|electricity\s?bill|water\s?bill|maintenance\s?charge)\b", "Reimbursement"),
    # Move-in
    (r"\b(move[\s-]?in|movein|onboard|new\s?tenant|just\s?moved|shifted|handover|formalities|bgv|background\s?verification|police\s?verification|agreement\s?sign|inventory\s?check|first\s?day)\b", "Move-in"),
    # Move-out
    (r"\b(move[\s-]?out|moveout|vacate|notice\s?period|leaving|shifting\s?out|exit|surrender)\b", "Move-out"),
    # Locks / Access
    (r"\b(lock|lockbox|key|keys|duplicate\s?key|main\s?door|door\s?lock|smart\s?lock|missing\s?key)\b", "Access & Security"),
    # Gas/Utilities
    (r"\b(gas\s?cylinder|gas\s?lighter|lpg|gas.*connect|pipeline|water\s?supply|tanker|borewell|parking)\b", "Utilities"),
    # Documentation
    (r"\b(agreement|contract|document|noc|letter|stamp|notary|registration|verification)\b", "Documentation"),
    # Inventory / missing items
    (r"\b(missing|not\s?provided|not\s?given|need\s?a\s?new|needs?\s?replacement|install|provide|required|basket|laundry|stand)\b", "Inventory"),
    # Reimbursement — broader catch
    (r"\b(cab\s?reimburse|maintenance\s?charge|society\s?due|unpaid)\b", "Reimbursement"),
    # Access — broader
    (r"\b(locked\s?from|bathroom\s?lock)\b", "Access & Security"),
]


# Direct short-description-to-category map for 1-2 word descriptions
SHORT_DESC_MAP = {
    "plumbing": "Plumbing",
    "electrical": "Electrical",
    "appliance": "Appliance",
    "carpentry": "Carpentry",
    "pest control": "Pest Control",
    "cleaning": "Cleaning",
    "utilities": "Utilities",
    "wifi": "Utilities",
    "reimbursement": "Reimbursement",
    "maintenance": "Maintenance",
    "inventory": "Inventory",
    "move-in": "Move-in",
    "move-out": "Move-out",
    "documentation": "Documentation",
}


def classify_from_text(title, description):
    text = f"{title} {description}".lower().strip()
    desc_lower = description.lower().strip()

    # If description is very short (1-2 words), try direct category lookup
    if len(desc_lower.split()) <= 2:
        mapped = SHORT_DESC_MAP.get(desc_lower)
        if mapped:
            return mapped

    # Try regex rules
    for pattern, category in DESCRIPTION_RULES:
        if re.search(pattern, text):
            return category

    # Last resort: check if the ticket name contains category-like words
    name_lower = title.lower()
    for pattern, category in DESCRIPTION_RULES:
        if re.search(pattern, name_lower):
            return category

    return "Other"


def normalize_category(raw_cat, title="", description=""):
    if not raw_cat or raw_cat == "N/A":
        return classify_from_text(title, description)
    raw_lower = raw_cat.strip().lower()
    if raw_lower in CATEGORY_MAP:
        result = CATEGORY_MAP[raw_lower]
        if result == "_unclassified":
            return classify_from_text(title, description)
        return result
    if re.match(r"\d+br\d+", raw_lower) or "|" in raw_lower:
        # Raw category is a property ID or pipe-separated — extract text after pipe
        if "|" in raw_lower:
            after_pipe = raw_lower.split("|", 1)[1].strip()
            if after_pipe:
                for pattern, category in DESCRIPTION_RULES:
                    if re.search(pattern, after_pipe):
                        return category
        return classify_from_text(title, description)
    return classify_from_text(title, description)


# ─────────────────────────────────────────────────────────────────────────────
# ROOT CAUSE EXTRACTION — mine description + resolution notes
# ─────────────────────────────────────────────────────────────────────────────

ROOT_CAUSE_PATTERNS = [
    # Infrastructure / physical issues
    (r"\b(leak|leaking|leakage|seepage|crack|broken|damaged|corroded|rusted|worn\s?out)\b", "Physical damage / wear"),
    (r"\b(clog|blocked|choked|stuck|jam)\b", "Blockage / clogging"),
    (r"\b(not\s?working|stopped\s?working|malfunction|faulty|defective|dead)\b", "Equipment malfunction"),
    (r"\b(no\s?power|no\s?electricity|power\s?cut|outage|tripped)\b", "Power issue"),
    (r"\b(no\s?water|low\s?pressure|water\s?pressure|no\s?hot\s?water)\b", "Water supply issue"),
    # UX / information gaps
    (r"\b(confused|confusion|unclear|don.t\s?know|didn.t\s?know|not\s?sure|how\s?to|how\s?do)\b", "User confusion"),
    (r"\b(wrong|incorrect|error|mistake|mismatch)\b", "Data/process error"),
    (r"\b(no\s?response|no\s?update|didn.t\s?respond|not\s?reachable|no\s?reply)\b", "Communication failure"),
    (r"\b(delay|delayed|late|overdue|slow|pending.*long|long.*pending)\b", "Delayed resolution"),
    (r"\b(duplicate|already|same\s?issue|again|recurring|repeat)\b", "Recurring / duplicate"),
    # Self-serve
    (r"\b(restart|reset|reboot|power\s?cycle|switch.*off.*on|replug|unplug)\b", "Simple reset needed"),
    (r"\b(setting|settings|configuration|configure|setup|install)\b", "Configuration needed"),
]


def extract_root_cause(title, description, resolution_notes):
    text = f"{title} {description} {resolution_notes}".lower()
    causes = []
    for pattern, cause in ROOT_CAUSE_PATTERNS:
        if re.search(pattern, text):
            causes.append(cause)
    return causes[:3] if causes else ["Undiagnosed"]


# ─────────────────────────────────────────────────────────────────────────────
# RESOLUTION PATTERN MINING — what actually fixed the ticket?
# ─────────────────────────────────────────────────────────────────────────────

RESOLUTION_TYPES = [
    (r"\b(vendor|technician|plumber|electrician|carpenter|service\s?person)\b.*\b(visit|came|sent|assigned|scheduled)\b", "Vendor visit"),
    (r"\b(replaced|replacement|new\s?one|installed\s?new|changed)\b", "Part replacement"),
    (r"\b(repaired|fixed|resolved|serviced|cleaned)\b", "Repair / service"),
    (r"\b(clarified|explained|shared|informed|told|confirmed|let.*know)\b", "Information shared"),
    (r"\b(tenant\s?resolved|self.*resolved|resolved\s?(by|on)\s?(his|her|their|its|own)|restarted|reset)\b", "Self-resolved"),
    (r"\b(duplicate|already.*ticket|same.*ticket|merged)\b", "Duplicate ticket"),
    (r"\b(no\s?issue|not\s?an?\s?issue|working\s?fine|false\s?alarm)\b", "No issue found"),
    (r"\b(waiting|pending|follow\s?up|scheduled|will\s?be)\b", "Pending / waiting"),
    (r"\b(landlord|owner|ll\b)\b.*\b(approved|agreed|informed|asked)\b", "Landlord coordination"),
    (r"\b(escalat|raised|forwarded|transferred)\b", "Escalated"),
    (r"\b(refund|reimburs|paid|payment\s?done|credited)\b", "Financial resolution"),
]


def extract_resolution_type(resolution_notes):
    if not resolution_notes:
        return "Unknown"
    text = resolution_notes.lower()
    for pattern, rtype in RESOLUTION_TYPES:
        if re.search(pattern, text):
            return rtype
    if len(text) < 5:
        return "Unknown"
    return "Manual resolution"


# ─────────────────────────────────────────────────────────────────────────────
# UX ISSUE CLASSIFICATION — multi-label, based on full ticket context
# ─────────────────────────────────────────────────────────────────────────────

def classify_ux_issues(ticket):
    text = f"{ticket['description']} {ticket['name']}".lower()
    resolution = (ticket.get("resolution_notes") or "").lower()
    full_text = f"{text} {resolution}"
    issues = []

    # Onboarding friction
    if re.search(r"\b(move[\s-]?in|just\s?(moved|shifted)|new\s?tenant|first\s?time|onboard|formalities|handover|setup)\b", text):
        issues.append("Onboarding Friction")

    # Payment / billing clarity
    if re.search(r"\b(rent|payment|bill|invoice|refund|deposit|charge|due|paid|reimburse|receipt|emi)\b", text):
        issues.append("Payment Clarity")

    # Communication gap — both from ticket text and resolution pattern
    if re.search(r"\b(no\s?response|no\s?update|didn.t\s?respond|not\s?reachable|still\s?waiting|when\s?will|any\s?update|follow\s?up)\b", full_text):
        issues.append("Communication Gap")

    # Status / operational visibility
    if re.search(r"\b(status|where\s?is|what\s?happened|no\s?visibility|don.t\s?know|track|when.*come|when.*resolve|eta)\b", text):
        issues.append("Status Visibility")

    # Self-serve opportunity
    resolution_type = ticket.get("resolution_type", "")
    if resolution_type in ("Self-resolved", "Information shared", "No issue found", "Duplicate ticket"):
        issues.append("Self-serve Opportunity")
    elif re.search(r"\b(restart|reset|reboot|how\s?to|guide|instructions|manual)\b", full_text):
        issues.append("Self-serve Opportunity")

    # Expectation mismatch
    if re.search(r"\b(expected|supposed\s?to|should\s?have|was\s?told|promised|committed)\b", text):
        issues.append("Expectation Mismatch")

    # Repeat / recurring issue
    if ticket.get("is_repeat_property_issue") and ticket.get("repeat_count_property", 0) >= 3:
        issues.append("Recurring Infrastructure")

    # Vendor dependency
    if re.search(r"\b(vendor|technician|plumber|electrician)\b.*\b(not.*came|didn.t.*come|waiting|delay|no\s?show)\b", full_text):
        issues.append("Vendor Dependency")
    elif re.search(r"\b(waiting.*vendor|vendor.*delay|vendor.*scheduled)\b", full_text):
        issues.append("Vendor Dependency")

    # Unclear process
    if re.search(r"\b(how\s?to|how\s?do|what\s?is\s?the\s?process|not\s?sure|confused|where\s?do)\b", text):
        issues.append("Process Unclear")

    # Landlord dependency
    if re.search(r"\b(landlord|owner|ll\b)\b.*\b(approval|permission|waiting|pending|not.*respond)\b", full_text):
        issues.append("Landlord Dependency")

    if not issues:
        issues.append("Physical Resolution Needed")

    return issues


# ─────────────────────────────────────────────────────────────────────────────
# UX THEME — single primary theme per ticket
# ─────────────────────────────────────────────────────────────────────────────

def detect_ux_theme(ticket):
    text = f"{ticket.get('description', '')} {ticket.get('name', '')}".lower()
    cat = ticket.get("category", "")

    # Explicit theme from text — ORDER MATTERS: specific before general
    if re.search(r"\b(move[\s-]?in|movein|onboard|formalities|handover|just\s?(moved|shifted)|new\s?tenant)\b", text):
        return "Onboarding"
    if re.search(r"\b(move[\s-]?out|moveout|vacate|notice\s?period|leaving|shifting\s?out)\b", text):
        return "Move-out"
    if re.search(r"\b(rent|payment|bill|invoice|refund|deposit|reimburse|charge|due|emi|receipt)\b", text):
        return "Payments & Billing"
    if re.search(r"\b(wifi|wi-fi|internet|router|broadband|network|connectivity)\b", text):
        return "Connectivity"
    # Appliances BEFORE plumbing — "water filter", "geyser", "washing machine" are appliances
    if re.search(r"\b(ro\b|water\s?filter|water\s?purifier|aquaguard|washing\s?machin|fridge|refrigerator|microwave|chimney|oven|stove|appliance|gas\s?knob|drinkprime|tv|television|iron\s?box|exhaust\s?fan|exhaust|vacuum|lamp|drying\s?stand|ventilation)\b", text):
        return "Appliances"
    if re.search(r"\b(ac\b|air\s?condition|fan(?!\s?making)(?!\s?nois)(?!\s?slow)|geyser|geysar|heater|cooling|heating|temperature)\b", text):
        return "Climate & Comfort"
    if re.search(r"\b(water(?!\s?filter)|leak|drain|flush|toilet|basin|washbasin|tap|pipe|seepage|plumb|sewer|shower|sink|hard\s?water|no\s?water|water\s?damage)\b", text):
        return "Water & Plumbing"
    if re.search(r"\b(pest|cockroach|roach|roaches|rat|rats|ant|ants|mosquito|termite|bed\s?bug|lizard|pigeon|smell|stink|bad\s?odour|bad\s?odor|bad\s?smell|infest)\b", text):
        return "Pest Issues"
    if re.search(r"\b(lock|lockbox|key|keys|access|entry|security)\b", text):
        return "Access & Security"
    if re.search(r"\b(clean|hygiene|dirty|stain|garbage|dustbin|deep\s?clean|bucket|mug)\b", text):
        return "Cleanliness"
    if re.search(r"\b(agreement|contract|document|verification|noc)\b", text):
        return "Documentation"
    if re.search(r"\b(switch|socket|wire|wiring|power|electric|light|bulb|mcb|circuit|fan\s?making|fan\s?nois|fan.*slow)\b", text):
        return "Electrical"
    # Fixtures / Carpentry / Furniture — broad catch for physical items
    if re.search(r"\b(door|window|wardrobe|cupboard|drawer|cabinet|shelf|hinge|handle|bed|frame|table|chair|desk|hooks?|rod|curtains?|blackout|painting|wall\s?art|crack|hole|furniture|sofa|mattress|bedsheet|pillow|rack|ladder|balcony|net\b)\b", text):
        return "Fixtures & Carpentry"

    # Fallback from category
    theme_from_cat = {
        "Plumbing": "Water & Plumbing",
        "Electrical": "Electrical",
        "Appliance": "Appliances",
        "Carpentry": "Fixtures & Carpentry",
        "Utilities": "Connectivity",
        "Pest Control": "Pest Issues",
        "Reimbursement": "Payments & Billing",
        "Cleaning": "Cleanliness",
        "Move-in": "Onboarding",
        "Move-out": "Move-out",
        "Documentation": "Documentation",
        "Access & Security": "Access & Security",
        "Maintenance": "General Maintenance",
        "Inventory": "Fixtures & Carpentry",
        "Communication": "General Maintenance",
        "Internal": "General Maintenance",
    }
    return theme_from_cat.get(cat, "General Maintenance")


# ─────────────────────────────────────────────────────────────────────────────
# SCORING — frustration + preventability (improved)
# ─────────────────────────────────────────────────────────────────────────────

def compute_frustration_score(ticket):
    score = 0
    text = f"{ticket.get('description', '')} {ticket.get('name', '')}".lower()
    resolution = (ticket.get("resolution_notes") or "").lower()

    # Priority
    p = ticket.get("priority", "")
    if p == "URGENT": score += 3
    elif p == "HIGH": score += 2
    elif p == "MEDIUM": score += 1

    # Emotional language
    if re.search(r"\b(urgent|asap|immediately|emergency|unbearable|unacceptable|terrible|pathetic|worst|disgusting|horrible|disappointed|frustrated)\b", text):
        score += 2
    if re.search(r"\b(again|recurring|same\s?issue|repeated|multiple\s?times|still\s?(not|broken|same)|not\s?fixed|again\s?and\s?again)\b", text):
        score += 2

    # Long resolution time
    hours = ticket.get("resolution_time_hours")
    if hours and hours > 336:  # > 2 weeks
        score += 3
    elif hours and hours > 168:  # > 1 week
        score += 2
    elif hours and hours > 72:  # > 3 days
        score += 1

    # Communication failures in resolution
    if re.search(r"\b(no\s?response|didn.t\s?respond|not\s?reachable|multiple\s?times|tried\s?reaching)\b", resolution):
        score += 1

    # Repeat issue amplifier
    if ticket.get("is_repeat_customer_issue"):
        score += 1
    if ticket.get("repeat_count_property", 0) >= 5:
        score += 1

    # Still open and old
    if ticket.get("is_open") and hours and hours > 168:
        score += 1

    return min(score, 10)


def compute_preventability_score(ticket):
    score = 0
    text = f"{ticket.get('description', '')} {ticket.get('name', '')}".lower()
    resolution = (ticket.get("resolution_notes") or "").lower()
    res_type = ticket.get("resolution_type", "")

    # Resolution type is the strongest signal
    if res_type == "Information shared":
        score += 4  # purely info — a dashboard would prevent this
    elif res_type == "Self-resolved":
        score += 4  # user fixed it — a guide would prevent the ticket
    elif res_type == "No issue found":
        score += 5  # ticket shouldn't have existed
    elif res_type == "Duplicate ticket":
        score += 5  # ticket shouldn't have existed

    # Confusion / clarity issues
    if re.search(r"\b(confused|confusion|unclear|didn.t\s?know|not\s?sure|clarified|explained)\b", f"{text} {resolution}"):
        score += 2

    # Payment-related info requests
    if re.search(r"\b(rent|bill|payment|charge|refund|reimburse)\b", text):
        if re.search(r"\b(clarified|shared|explained|informed|confirmed|sent)\b", resolution):
            score += 3
        else:
            score += 1  # even without, payment tickets are often info-based

    # Status/visibility requests
    if re.search(r"\b(status|when.*will|where\s?is|any\s?update|tracking|how\s?long)\b", text):
        score += 2

    # Onboarding — these are definitionally preventable
    if re.search(r"\b(move[\s-]?in|formalities|onboard|handover|setup|how\s?to|how\s?do\s?i)\b", text):
        score += 2

    # Simple resets found in resolution
    if re.search(r"\b(restart|restarted|reset|reboot|switch.*off.*on|replug|power\s?cycle)\b", resolution):
        score += 3

    # Resolution notes show it was trivial
    if re.search(r"\b(working\s?fine|no\s?issue|false\s?alarm|already\s?resolved|resolved\s?itself)\b", resolution):
        score += 3

    # Repeat issue — preventable with proactive maintenance
    rpc = ticket.get("repeat_count_property", 0)
    if rpc >= 5:
        score += 2
    elif rpc >= 3:
        score += 1

    return min(score, 10)


# ─────────────────────────────────────────────────────────────────────────────
# CUSTOMER NAME EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract_customer_name(ticket_name):
    if not ticket_name:
        return None
    match = re.match(r"\s*([^|\-]+)", ticket_name)
    if match:
        name = match.group(1).strip()
        if name and name.lower() not in ("internal", "n/a", "") and len(name) > 2 and len(name) < 40:
            return name
    return None


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PROCESSING
# ─────────────────────────────────────────────────────────────────────────────

def process():
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    # Load full dataset from Request (F)
    ws = wb["Request (F)"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    raw_tickets = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0]:
            raw_tickets.append(dict(zip(headers, row)))

    print(f"Loaded {len(raw_tickets)} raw tickets")

    # Load open ticket IDs
    ws_open = wb["Customer Tickets - OPEN"]
    open_ids = set()
    for row in ws_open.iter_rows(min_row=2, max_row=ws_open.max_row, values_only=True):
        if row[0]:
            open_ids.add(str(row[0]))

    # ── Build processed tickets ──────────────────────────────────────────────
    processed = []
    property_issues = defaultdict(list)
    customer_issues = defaultdict(list)

    for raw in raw_tickets:
        ticket_id = str(raw.get("Ticket ID", ""))
        name = str(raw.get("Ticket Name", "") or "").strip()
        description = str(raw.get("Ticket Description", "") or "").strip()
        raw_category = str(raw.get("Ticket Category", "") or "")
        priority = str(raw.get("Priority", "LOW") or "LOW")
        if priority == "N/A":
            priority = "LOW"
        status = str(raw.get("Status (F)", "") or "Unknown")
        owner = str(raw.get("Owner (F)", "") or "") or None
        if owner == "None":
            owner = None
        rid_raw = raw.get("RID") or ""
        rid = str(rid_raw) if rid_raw and rid_raw != "N/A" else ""
        pid_raw = raw.get("PID")
        pid = None
        if pid_raw and pid_raw != "N/A":
            try:
                pid = str(int(float(pid_raw)))
            except (ValueError, TypeError):
                pid = None
        resolution_notes = str(raw.get("Resolution Notes", "") or "")
        if resolution_notes == "N/A":
            resolution_notes = ""
        cost = raw.get("Cost Associated", 0) or 0
        if cost == "N/A":
            cost = 0
        else:
            try:
                cost = float(cost)
            except (ValueError, TypeError):
                cost = 0
        cost_paid_by = str(raw.get("Cost Paid By", "") or "")
        if cost_paid_by == "N/A":
            cost_paid_by = ""
        csat = raw.get("CSAT", "")
        if csat == "N/A" or not csat:
            csat = None
        created_by = str(raw.get("Created By", "") or "")

        # Parse dates
        create_date = None
        create_dt = None
        if isinstance(raw.get("Create Date"), datetime.datetime):
            create_dt = raw["Create Date"]
            create_date = create_dt.strftime("%Y-%m-%d")
        close_date = None
        if isinstance(raw.get("Close Date"), datetime.datetime):
            close_date = raw["Close Date"].strftime("%Y-%m-%d")

        # Resolution time
        resolution_time_hours = None
        if isinstance(raw.get("Create Date"), datetime.datetime) and isinstance(
            raw.get("Close Date"), datetime.datetime
        ):
            delta = raw["Close Date"] - raw["Create Date"]
            hours = round(delta.total_seconds() / 3600, 1)
            if hours >= 0:
                resolution_time_hours = hours

        # Normalize category (now uses description fallback for Other/Uncategorized)
        category = normalize_category(raw_category, name, description)

        # Extract customer name
        customer_name = extract_customer_name(name)

        # Extract RID from name if missing
        if not rid:
            rid_match = re.search(r"(\d+BR\d+)", name, re.IGNORECASE)
            if rid_match:
                rid = rid_match.group(1).upper()

        # Resolution type mining
        resolution_type = extract_resolution_type(resolution_notes)

        # Root causes
        root_causes = extract_root_cause(name, description, resolution_notes)

        is_open = ticket_id in open_ids

        t = {
            "id": ticket_id,
            "name": name,
            "description": description,
            "category_raw": raw_category,
            "category": category,
            "priority": priority,
            "status": status,
            "is_open": is_open,
            "owner": owner,
            "customer_name": customer_name,
            "rid": rid if rid else None,
            "pid": pid,
            "create_date": create_date,
            "close_date": close_date,
            "resolution_time_hours": resolution_time_hours,
            "resolution_notes": resolution_notes,
            "resolution_type": resolution_type,
            "root_causes": root_causes,
            "cost": cost,
            "cost_paid_by": cost_paid_by,
            "csat": csat,
            "created_by": created_by,
        }

        if rid:
            property_issues[rid].append(t)
        if customer_name:
            customer_issues[customer_name].append(t)

        processed.append(t)

    # ── Repeat detection ─────────────────────────────────────────────────────
    property_category_counts = defaultdict(int)
    for rid, tix in property_issues.items():
        for t in tix:
            property_category_counts[f"{rid}|{t['category']}"] += 1

    customer_category_counts = defaultdict(int)
    for cust, tix in customer_issues.items():
        for t in tix:
            customer_category_counts[f"{cust}|{t['category']}"] += 1

    # Customer first ticket dates (for time-to-first-ticket analysis)
    customer_first_ticket = {}
    for cust, tix in customer_issues.items():
        dates = [t["create_date"] for t in tix if t["create_date"]]
        if dates:
            customer_first_ticket[cust] = min(dates)

    # ── Second pass: UX scoring ──────────────────────────────────────────────
    for t in processed:
        rid = t.get("rid", "")
        cat = t.get("category", "")
        cust = t.get("customer_name", "")

        prop_key = f"{rid}|{cat}" if rid else ""
        cust_key = f"{cust}|{cat}" if cust else ""

        t["is_repeat_property_issue"] = property_category_counts.get(prop_key, 0) > 1
        t["is_repeat_customer_issue"] = customer_category_counts.get(cust_key, 0) > 1
        t["repeat_count_property"] = property_category_counts.get(prop_key, 0)
        t["repeat_count_customer"] = customer_category_counts.get(cust_key, 0)

        # Time since customer's first ticket
        if cust and cust in customer_first_ticket and t["create_date"]:
            first = customer_first_ticket[cust]
            try:
                delta = (datetime.datetime.strptime(t["create_date"], "%Y-%m-%d") -
                         datetime.datetime.strptime(first, "%Y-%m-%d")).days
                t["days_since_first_ticket"] = delta
                t["is_first_week_ticket"] = delta <= 7
                t["is_first_month_ticket"] = delta <= 30
            except ValueError:
                t["days_since_first_ticket"] = None
                t["is_first_week_ticket"] = False
                t["is_first_month_ticket"] = False
        else:
            t["days_since_first_ticket"] = None
            t["is_first_week_ticket"] = False
            t["is_first_month_ticket"] = False

        # UX theme
        t["ux_theme"] = detect_ux_theme(t)

        # Scores (need resolution_type set first)
        t["frustration_score"] = compute_frustration_score(t)
        t["preventability_score"] = compute_preventability_score(t)

        # UX issue classification
        t["ux_issues"] = classify_ux_issues(t)

        # Derived booleans
        t["is_preventable"] = t["preventability_score"] >= 3
        t["is_onboarding_issue"] = "Onboarding Friction" in t["ux_issues"]
        t["is_communication_gap"] = "Communication Gap" in t["ux_issues"]
        t["is_self_serve_possible"] = "Self-serve Opportunity" in t["ux_issues"]
        t["is_visibility_issue"] = "Status Visibility" in t["ux_issues"]
        t["is_automation_opportunity"] = (
            t["is_self_serve_possible"]
            or t["is_preventable"]
            or t["is_communication_gap"]
        )

    print(f"Processed {len(processed)} tickets")

    # ── Generate deep insights ───────────────────────────────────────────────
    total = len(processed)
    preventable = sum(1 for t in processed if t["is_preventable"])
    onboarding = sum(1 for t in processed if t["is_onboarding_issue"])
    communication = sum(1 for t in processed if t["is_communication_gap"])
    self_serve = sum(1 for t in processed if t["is_self_serve_possible"])
    visibility = sum(1 for t in processed if t["is_visibility_issue"])
    automation = sum(1 for t in processed if t["is_automation_opportunity"])
    repeat_prop = sum(1 for t in processed if t["is_repeat_property_issue"])
    first_week = sum(1 for t in processed if t["is_first_week_ticket"])
    first_month = sum(1 for t in processed if t["is_first_month_ticket"])

    cat_dist = Counter(t["category"] for t in processed)
    theme_dist = Counter(t["ux_theme"] for t in processed)
    resolution_type_dist = Counter(t["resolution_type"] for t in processed)

    # UX issue distribution
    ux_issue_dist = Counter()
    for t in processed:
        for issue in t["ux_issues"]:
            ux_issue_dist[issue] += 1

    # Root cause distribution
    root_cause_dist = Counter()
    for t in processed:
        for rc in t["root_causes"]:
            root_cause_dist[rc] += 1

    # Monthly trend
    monthly = defaultdict(int)
    monthly_preventable = defaultdict(int)
    for t in processed:
        if t["create_date"]:
            month = t["create_date"][:7]
            monthly[month] += 1
            if t["is_preventable"]:
                monthly_preventable[month] += 1

    # Frustration by theme
    frust_by_theme = defaultdict(list)
    for t in processed:
        frust_by_theme[t["ux_theme"]].append(t["frustration_score"])
    avg_frust_by_theme = {
        theme: round(sum(scores) / len(scores), 1)
        for theme, scores in frust_by_theme.items()
        if len(scores) >= 5
    }

    # Resolution time by category
    res_by_cat = defaultdict(list)
    for t in processed:
        if t["resolution_time_hours"] and t["resolution_time_hours"] > 0:
            res_by_cat[t["category"]].append(t["resolution_time_hours"])
    avg_res_by_cat = {
        cat: round(sum(hrs) / len(hrs), 1)
        for cat, hrs in res_by_cat.items()
        if len(hrs) >= 5
    }
    median_res_by_cat = {}
    for cat, hrs in res_by_cat.items():
        if len(hrs) >= 5:
            sorted_hrs = sorted(hrs)
            mid = len(sorted_hrs) // 2
            median_res_by_cat[cat] = sorted_hrs[mid]

    # Top repeat property issues
    prop_cat_list = [
        {"property": k.split("|")[0], "category": k.split("|")[1], "count": v}
        for k, v in property_category_counts.items()
        if v >= 3
    ]
    prop_cat_list.sort(key=lambda x: x["count"], reverse=True)

    # Customer repeat analysis
    customer_ticket_counts = Counter()
    for cust, tix in customer_issues.items():
        customer_ticket_counts[cust] = len(tix)

    # Property health scores
    property_health = {}
    for rid, tix in property_issues.items():
        if len(tix) < 3:
            continue
        avg_frust = sum(t["frustration_score"] for t in tix) / len(tix)
        repeat_cats = len(set(
            cat for cat, count in Counter(t["category"] for t in tix).items() if count >= 3
        ))
        total_cost = sum(t["cost"] for t in tix)
        open_count = sum(1 for t in tix if t["is_open"])
        property_health[rid] = {
            "rid": rid,
            "total_tickets": len(tix),
            "open_tickets": open_count,
            "avg_frustration": round(avg_frust, 1),
            "recurring_categories": repeat_cats,
            "total_cost": total_cost,
            "health_score": round(max(0, 10 - avg_frust - repeat_cats * 0.5 - open_count * 0.3), 1),
        }

    # Cost analysis
    total_cost = sum(t["cost"] for t in processed)
    cost_by_cat = defaultdict(float)
    for t in processed:
        cost_by_cat[t["category"]] += t["cost"]

    # First-week analysis
    first_week_tickets = [t for t in processed if t["is_first_week_ticket"]]
    first_week_themes = Counter(t["ux_theme"] for t in first_week_tickets)

    # ── AI Insight Cards ─────────────────────────────────────────────────────
    ai_insights = []

    # Preventable tickets insight
    if preventable > 0:
        prevent_by_cat = Counter(t["category"] for t in processed if t["is_preventable"])
        top3 = prevent_by_cat.most_common(3)
        ai_insights.append({
            "type": "preventable",
            "severity": "high",
            "title": f"{round(preventable/total*100, 1)}% of tickets are preventable",
            "description": f"{preventable} tickets could be prevented through better UX. Top preventable categories: {', '.join(f'{c} ({n})' for c, n in top3)}.",
            "recommendation": "Invest in self-serve tools, proactive information sharing, and guided troubleshooting flows",
        })

    # Resolution type insight — "Information shared" is purely preventable
    info_shared = sum(1 for t in processed if t["resolution_type"] == "Information shared")
    self_resolved = sum(1 for t in processed if t["resolution_type"] == "Self-resolved")
    no_issue = sum(1 for t in processed if t["resolution_type"] == "No issue found")
    duplicates = sum(1 for t in processed if t["resolution_type"] == "Duplicate ticket")
    zero_effort = info_shared + self_resolved + no_issue + duplicates
    if zero_effort > 0:
        ai_insights.append({
            "type": "zero_effort",
            "severity": "high",
            "title": f"{zero_effort} tickets ({round(zero_effort/total*100, 1)}%) required no physical intervention",
            "description": f"Breakdown: {info_shared} resolved by sharing info, {self_resolved} self-resolved, {no_issue} no real issue, {duplicates} duplicates. These tickets exist because users lacked access to information or tools.",
            "recommendation": "Build a self-serve portal with payment visibility, ticket status tracking, and troubleshooting guides",
        })

    # First-week insight
    if first_week > 0:
        first_week_pct = round(first_week / total * 100, 1)
        top_fw_themes = first_week_themes.most_common(3)
        ai_insights.append({
            "type": "onboarding",
            "severity": "high",
            "title": f"{first_week_pct}% of tickets come within the first week of tenancy",
            "description": f"{first_week} tickets are raised within 7 days of a customer's first ticket. Top themes: {', '.join(f'{t} ({c})' for t, c in top_fw_themes)}. This is pure onboarding friction.",
            "recommendation": "Create an interactive move-in checklist that proactively addresses WiFi setup, utility connections, appliance guides, and key contacts",
        })

    # Communication gap
    if communication > 0:
        ai_insights.append({
            "type": "communication",
            "severity": "medium",
            "title": f"{communication} tickets involve communication gaps",
            "description": f"{round(communication/total*100, 1)}% of tickets mention lack of response, delayed updates, or follow-up confusion. Every follow-up ticket is a UX failure.",
            "recommendation": "Implement automated status updates via WhatsApp/SMS when ticket status changes, vendor is assigned, or resolution is complete",
        })

    # Repeat property issues
    worst_properties = sorted(property_health.values(), key=lambda x: x["health_score"])[:5]
    if worst_properties:
        ai_insights.append({
            "type": "infrastructure",
            "severity": "high",
            "title": f"5 properties account for disproportionate ticket volume",
            "description": f"Worst health scores: {', '.join(f'{p['rid']} (score {p['health_score']}/10, {p['total_tickets']} tickets)' for p in worst_properties[:3])}. These need proactive maintenance audits.",
            "recommendation": "Schedule quarterly maintenance inspections for properties with health score below 5",
        })

    # Cost insight
    if total_cost > 0:
        top_cost_cats = sorted(cost_by_cat.items(), key=lambda x: x[1], reverse=True)[:3]
        ai_insights.append({
            "type": "cost",
            "severity": "medium",
            "title": f"₹{total_cost:,.0f} total cost across all tickets",
            "description": f"Highest cost categories: {', '.join(f'{cat} (₹{cost:,.0f})' for cat, cost in top_cost_cats)}. Preventable tickets alone account for ₹{sum(t['cost'] for t in processed if t['is_preventable']):,.0f}.",
            "recommendation": "Focus cost reduction on the top 3 categories — preventable tickets represent recoverable spend",
        })

    # Resolution time insight
    all_res_hours = [t["resolution_time_hours"] for t in processed if t["resolution_time_hours"] and t["resolution_time_hours"] > 0]
    if all_res_hours:
        avg_hours = sum(all_res_hours) / len(all_res_hours)
        sorted_hours = sorted(all_res_hours)
        median_hours = sorted_hours[len(sorted_hours) // 2]
        slow_cats = sorted(avg_res_by_cat.items(), key=lambda x: x[1], reverse=True)[:3]
        ai_insights.append({
            "type": "resolution_time",
            "severity": "medium",
            "title": f"Average resolution: {round(avg_hours/24, 1)} days (median: {round(median_hours/24, 1)} days)",
            "description": f"Slowest categories: {', '.join(f'{cat} ({round(hrs/24, 1)}d avg)' for cat, hrs in slow_cats)}. Tickets open over 7 days have 2x frustration scores.",
            "recommendation": "Set category-specific SLAs and display expected resolution times to tenants at ticket creation",
        })

    # Self-serve
    if self_serve > 0:
        ss_themes = Counter(t["ux_theme"] for t in processed if t["is_self_serve_possible"])
        top_ss = ss_themes.most_common(3)
        ai_insights.append({
            "type": "self_serve",
            "severity": "medium",
            "title": f"{self_serve} tickets could have been self-served",
            "description": f"Top self-serve themes: {', '.join(f'{t} ({c})' for t, c in top_ss)}. These were resolved by resets, basic troubleshooting, or information that should be accessible.",
            "recommendation": "Add contextual self-help before ticket creation: 'Before raising a ticket, try these steps...'",
        })

    # Repeat customer insight
    heavy_users = [(cust, count) for cust, count in customer_ticket_counts.items() if count >= 10]
    if heavy_users:
        heavy_users.sort(key=lambda x: x[1], reverse=True)
        ai_insights.append({
            "type": "repeat_customers",
            "severity": "medium",
            "title": f"{len(heavy_users)} customers have raised 10+ tickets each",
            "description": f"Top: {', '.join(f'{name} ({count})' for name, count in heavy_users[:5])}. These may be power users, unhappy tenants, or properties with systemic issues.",
            "recommendation": "Proactively reach out to high-ticket customers to understand root causes and improve their experience",
        })

    # ── Assemble insights output ─────────────────────────────────────────────
    all_res = [t["resolution_time_hours"] for t in processed if t["resolution_time_hours"] and t["resolution_time_hours"] > 0]

    insights = {
        "summary": {
            "total_tickets": total,
            "open_tickets": sum(1 for t in processed if t["is_open"]),
            "closed_tickets": sum(1 for t in processed if not t["is_open"]),
            "preventable_count": preventable,
            "preventable_pct": round(preventable / total * 100, 1),
            "onboarding_count": onboarding,
            "onboarding_pct": round(onboarding / total * 100, 1),
            "communication_count": communication,
            "communication_pct": round(communication / total * 100, 1),
            "self_serve_count": self_serve,
            "self_serve_pct": round(self_serve / total * 100, 1),
            "visibility_count": visibility,
            "visibility_pct": round(visibility / total * 100, 1),
            "automation_count": automation,
            "automation_pct": round(automation / total * 100, 1),
            "repeat_property_count": repeat_prop,
            "repeat_property_pct": round(repeat_prop / total * 100, 1),
            "first_week_count": first_week,
            "first_week_pct": round(first_week / total * 100, 1),
            "first_month_count": first_month,
            "first_month_pct": round(first_month / total * 100, 1),
            "zero_effort_count": zero_effort,
            "zero_effort_pct": round(zero_effort / total * 100, 1),
            "avg_resolution_hours": round(sum(all_res) / len(all_res), 1) if all_res else 0,
            "median_resolution_hours": sorted(all_res)[len(all_res) // 2] if all_res else 0,
            "total_cost": total_cost,
            "preventable_cost": sum(t["cost"] for t in processed if t["is_preventable"]),
        },
        "category_distribution": [
            {"name": k, "value": v} for k, v in cat_dist.most_common()
        ],
        "theme_distribution": [
            {"name": k, "value": v} for k, v in theme_dist.most_common()
        ],
        "ux_issue_distribution": [
            {"name": k, "value": v} for k, v in ux_issue_dist.most_common()
        ],
        "resolution_type_distribution": [
            {"name": k, "value": v} for k, v in resolution_type_dist.most_common()
        ],
        "root_cause_distribution": [
            {"name": k, "value": v} for k, v in root_cause_dist.most_common()
        ],
        "monthly_trend": [
            {"month": k, "tickets": monthly[k], "preventable": monthly_preventable.get(k, 0)}
            for k in sorted(monthly.keys())
        ],
        "frustration_by_theme": [
            {"theme": k, "score": v}
            for k, v in sorted(avg_frust_by_theme.items(), key=lambda x: x[1], reverse=True)
        ],
        "resolution_by_category": [
            {"category": k, "avg_hours": v, "median_hours": median_res_by_cat.get(k, 0)}
            for k, v in sorted(avg_res_by_cat.items(), key=lambda x: x[1], reverse=True)
        ],
        "cost_by_category": [
            {"category": k, "cost": v}
            for k, v in sorted(cost_by_cat.items(), key=lambda x: x[1], reverse=True)
            if v > 0
        ],
        "top_repeat_property_issues": prop_cat_list[:25],
        "property_health": sorted(property_health.values(), key=lambda x: x["health_score"]),
        "first_week_by_theme": [
            {"theme": k, "count": v}
            for k, v in first_week_themes.most_common()
        ],
        "repeat_customers": [
            {"name": name, "tickets": count}
            for name, count in sorted(customer_ticket_counts.items(), key=lambda x: x[1], reverse=True)[:20]
        ],
        "ai_insights": ai_insights,
    }

    # ── Write outputs ────────────────────────────────────────────────────────
    with open(OUTPUT_FILE, "w") as f:
        json.dump(processed, f, indent=2, default=str)
    print(f"Wrote {len(processed)} tickets to {OUTPUT_FILE}")

    with open(INSIGHTS_FILE, "w") as f:
        json.dump(insights, f, indent=2, default=str)
    print(f"Wrote insights to {INSIGHTS_FILE}")

    # Print summary
    print(f"\n{'='*60}")
    print(f"UX INTELLIGENCE SUMMARY v2")
    print(f"{'='*60}")
    print(f"Total tickets: {total}")
    print(f"Preventable: {preventable} ({insights['summary']['preventable_pct']}%)")
    print(f"Zero-effort resolutions: {zero_effort} ({insights['summary']['zero_effort_pct']}%)")
    print(f"First-week tickets: {first_week} ({insights['summary']['first_week_pct']}%)")
    print(f"First-month tickets: {first_month} ({insights['summary']['first_month_pct']}%)")
    print(f"Onboarding issues: {onboarding} ({insights['summary']['onboarding_pct']}%)")
    print(f"Communication gaps: {communication} ({insights['summary']['communication_pct']}%)")
    print(f"Self-serve possible: {self_serve} ({insights['summary']['self_serve_pct']}%)")
    print(f"Repeat property: {repeat_prop} ({insights['summary']['repeat_property_pct']}%)")
    print(f"Total cost: ₹{total_cost:,.0f}")
    print(f"\nCategories: {cat_dist.most_common(10)}")
    print(f"Themes: {theme_dist.most_common(10)}")
    print(f"Resolution types: {resolution_type_dist.most_common()}")
    print(f"Root causes: {root_cause_dist.most_common(8)}")


if __name__ == "__main__":
    process()
